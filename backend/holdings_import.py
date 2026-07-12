# -*- coding: utf-8 -*-
"""Preview parser for user-exported fund holding statements.

The parser deliberately never authenticates with a fund platform and never
persists the original file. It turns common CSV/XLSX export columns into
editable holding candidates, then the existing confirmed-holdings endpoint
performs the only write.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
import unicodedata
from typing import Any

from openpyxl import load_workbook


_MAX_ROWS = 1500
_HEADER_ALIASES = {
    "code": ("基金代码", "产品代码", "证券代码", "代码", "fundcode", "code"),
    "name": ("基金名称", "产品名称", "证券名称", "名称", "fundname", "name"),
    "amount": ("持有金额", "持仓金额", "持仓市值", "持有市值", "参考市值", "市值", "市值(元)", "amount", "marketvalue"),
    "cost": ("持仓成本", "成本金额", "总成本", "累计投入", "本金", "cost", "costamount"),
    "cost_price": ("持仓成本价", "成本价", "平均成本", "costprice"),
    "yesterday_profit": ("昨日收益", "昨日盈亏", "当日收益", "日收益", "昨日收益(元)", "dailyprofit"),
    "profit": ("持仓收益", "持有收益", "累计收益", "持仓盈亏", "累计盈亏", "收益金额", "profit"),
    "profit_rate": ("持仓收益率", "持有收益率", "累计收益率", "持仓盈亏率", "收益率", "profitrate"),
    "shares": ("持有份额", "持仓份额", "可用份额", "确认份额", "持有数量", "份额", "shares", "quantity"),
}


def _normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    return re.sub(r"[\s_\-()/\\\[\]{}:：]+", "", text)


def _text(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def _number(value: Any) -> float | None:
    text = _text(value)
    if not text or text in {"-", "--", "N/A", "n/a"}:
        return None
    text = text.replace(",", "").replace("¥", "").replace("￥", "").replace("元", "")
    text = re.sub(r"\s+", "", text)
    multiplier = 1.0
    if text.endswith("万"):
        multiplier = 10_000.0
        text = text[:-1]
    elif text.endswith("亿"):
        multiplier = 100_000_000.0
        text = text[:-1]
    text = text.rstrip("%")
    try:
        return float(text) * multiplier
    except ValueError:
        return None


def _clean_code(value: Any) -> str:
    code = _text(value)
    if re.fullmatch(r"\d+\.0", code):
        code = code[:-2]
    if code.isdigit() and len(code) < 6:
        return code.zfill(6)
    return code


def _find_mapping(headers: list[str]) -> dict[str, str | None]:
    normalized_headers = {_normalize(header): header for header in headers if header}
    mapping = {}
    for field, aliases in _HEADER_ALIASES.items():
        mapping[field] = next(
            (normalized_headers.get(_normalize(alias)) for alias in aliases if _normalize(alias) in normalized_headers),
            None,
        )
    return mapping


def _read_csv(data: bytes) -> tuple[list[str], list[dict[str, Any]], str, str]:
    text = None
    encoding = ""
    for candidate in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = data.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("账单编码无法识别，请导出 UTF-8 或 GBK/GB18030 格式的 CSV")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [_text(header) for header in (reader.fieldnames or [])]
    if not headers:
        raise ValueError("未识别到持仓账单表头")
    rows = []
    for row in reader:
        cleaned = {_text(key): _text(value) for key, value in row.items()}
        if any(cleaned.values()):
            rows.append(cleaned)
        if len(rows) > _MAX_ROWS:
            raise ValueError(f"单次最多预览 {_MAX_ROWS} 条持仓，请拆分账单后导入")
    delimiter_label = {",": "逗号", ";": "分号", "\t": "制表符"}.get(delimiter, delimiter)
    return headers, rows, f"CSV · {delimiter_label}", encoding


def _read_xlsx(data: bytes) -> tuple[list[str], list[dict[str, Any]], str, str]:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Excel 持仓账单无法读取:{exc}") from exc
    try:
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        header_values = next(values, None)
        while header_values is not None and not any(_text(value) for value in header_values):
            header_values = next(values, None)
        headers = [_text(value) for value in (header_values or [])]
        if not headers or not any(headers):
            raise ValueError("未识别到 Excel 持仓账单表头")
        rows = []
        for row_values in values:
            row = {
                headers[index]: row_values[index] if index < len(row_values) else ""
                for index in range(len(headers))
                if headers[index]
            }
            if any(_text(value) for value in row.values()):
                rows.append(row)
            if len(rows) > _MAX_ROWS:
                raise ValueError(f"单次最多预览 {_MAX_ROWS} 条持仓，请拆分账单后导入")
        return headers, rows, "Excel", "xlsx"
    finally:
        workbook.close()


def _read_statement(data: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]], str, str]:
    suffix = str(filename or "").lower().rsplit(".", 1)[-1]
    if suffix == "csv":
        return _read_csv(data)
    if suffix == "xlsx":
        return _read_xlsx(data)
    raise ValueError("请上传 CSV 或 XLSX 格式的持仓账单")


def _value(row: dict[str, Any], mapping: dict[str, str | None], field: str) -> Any:
    header = mapping.get(field)
    return row.get(header) if header else ""


def _template(filename: str) -> dict[str, str]:
    normalized_filename = _normalize(filename)
    if "天天基金" in str(filename or "") or "1234567" in normalized_filename:
        return {
            "id": "tiantian_fund",
            "label": "天天基金持仓导出",
            "source": "tiantian_fund_export",
            "detection": "文件名",
        }
    return {
        "id": "fund_holding_statement",
        "label": "基金持仓账单",
        "source": "holdings_file_import",
        "detection": "通用字段映射",
    }


def parse_holdings_file(data: bytes, *, filename: str = "") -> dict:
    """Produce editable fund-holding candidates without saving the uploaded file."""
    if not data:
        raise ValueError("持仓账单为空")
    headers, rows, file_format, encoding = _read_statement(data, filename)
    mapping = _find_mapping(headers)
    if not mapping.get("code"):
        raise ValueError("账单缺少基金代码列，无法安全识别持仓")

    template = _template(filename)
    candidates = []
    errors = []
    warnings = []
    seen_codes: dict[str, int] = {}
    for index, row in enumerate(rows, start=2):
        raw_code = _text(_value(row, mapping, "code"))
        code = _clean_code(raw_code)
        name = _text(_value(row, mapping, "name"))
        amount = _number(_value(row, mapping, "amount"))
        cost = _number(_value(row, mapping, "cost"))
        cost_price = _number(_value(row, mapping, "cost_price"))
        yesterday_profit = _number(_value(row, mapping, "yesterday_profit"))
        profit = _number(_value(row, mapping, "profit"))
        profit_rate = _number(_value(row, mapping, "profit_rate"))
        shares = _number(_value(row, mapping, "shares"))
        row_errors = []
        if not re.fullmatch(r"\d{6}", code):
            row_errors.append("基金代码必须是 6 位数字")
        if amount is not None and amount < 0:
            row_errors.append("持有金额不能小于 0")
        if cost is not None and cost < 0:
            row_errors.append("持仓成本不能小于 0")
        if shares is not None and shares < 0:
            row_errors.append("持有份额不能小于 0")
        if code and code in seen_codes:
            row_errors.append(f"与第 {seen_codes[code]} 行基金代码重复，避免覆盖持仓已排除")
        if row_errors:
            errors.append({"row": index, "message": "；".join(row_errors)})
            continue

        if raw_code != code:
            warnings.append(f"第 {index} 行基金代码已补齐前导零为 {code}，请确认。")
        if not name:
            warnings.append(f"第 {index} 行缺少基金名称，保存前请补充或核对。")
        if cost is None and cost_price is not None and shares is not None:
            cost = cost_price * shares
            warnings.append(f"第 {index} 行未提供成本金额，已由成本价乘以份额计算；请确认。")
        if amount is None and shares is None:
            warnings.append(f"第 {index} 行缺少持有金额和份额，组合配置分析不会计入该行。")

        candidates.append({
            "asset_type": "fund",
            "market": "基金",
            "code": code,
            "name": name,
            "amount": round(amount, 8) if amount is not None else None,
            "cost": round(cost, 8) if cost is not None else None,
            "yesterday_profit": round(yesterday_profit, 8) if yesterday_profit is not None else None,
            "profit": round(profit, 8) if profit is not None else None,
            "profit_rate": round(profit_rate, 8) if profit_rate is not None else None,
            "shares": round(shares, 8) if shares is not None else None,
            "source": template["source"],
            "raw_text": "",
            "csv_row": index,
        })
        seen_codes[code] = index

    if not candidates:
        raise ValueError("账单中没有可导入的有效基金持仓，请处理预览错误后重试")
    if errors:
        warnings.append(f"有 {len(errors)} 行无法安全导入，已从待确认列表中排除。")
    return {
        "source": "用户上传持仓账单（仅预览，未写入持仓）",
        "filename": str(filename or "")[:255],
        "file_sha256": hashlib.sha256(data).hexdigest(),
        "format": file_format,
        "encoding": encoding,
        "template": template,
        "headers": headers,
        "mapping": mapping,
        "candidates": candidates,
        "errors": errors,
        "warnings": warnings,
        "privacy": "原始持仓账单不会保存；确认后仅保存你核对过的持仓字段和来源。",
    }
