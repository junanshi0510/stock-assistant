# -*- coding: utf-8 -*-
"""Preview parser for user-exported transaction statements.

The parser never authenticates with a broker or fund platform, and it never
stores the original file.  It only turns confirmed transaction rows into an
editable preview; the existing confirmation endpoint remains the only write.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
import unicodedata
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook


_MAX_ROWS = 1500
_HEADER_ALIASES = {
    "trade_date": (
        "确认日期", "确认时间", "成交日期", "交易日期", "日期", "发生日期", "申请日期",
        "trade_date", "date",
    ),
    "code": ("证券代码", "基金代码", "基金编号", "股票代码", "代码", "产品代码", "symbol", "code", "ticker"),
    "name": ("证券名称", "基金名称", "股票名称", "名称", "产品名称", "name"),
    "trade_type": ("买卖标志", "交易方向", "业务名称", "业务类型", "交易类型", "方向", "操作", "trade_type", "side", "action"),
    "shares": ("成交数量", "成交份额", "确认份额", "确认份额(份)", "赎回份额", "数量", "份额", "股数", "shares", "quantity", "qty"),
    "unit_price": ("确认净值", "成交价格", "成交单价", "成交净值", "单位净值", "单价", "价格", "unit_price", "price"),
    "gross_amount": ("确认金额", "成交金额", "申请金额", "赎回金额", "发生金额", "金额", "交易金额", "amount", "gross_amount"),
    "fee": ("手续费", "申购费", "赎回费", "交易费用", "费用", "佣金", "fee", "commission"),
    "market": ("市场", "交易市场", "market"),
    "asset_type": ("资产类型", "产品类型", "asset_type", "asset"),
    "note": ("备注", "说明", "摘要", "note", "memo"),
    "status": ("确认状态", "交易状态", "状态", "处理状态", "status"),
}

_TYPE_MAP = {
    "买入": "buy",
    "买": "buy",
    "申购": "buy",
    "认购": "buy",
    "定投": "buy",
    "定投申购": "buy",
    "追加申购": "buy",
    "卖出": "sell",
    "卖": "sell",
    "赎回": "sell",
    "期初持仓": "opening",
    "期初": "opening",
    "opening": "opening",
    "buy": "buy",
    "sell": "sell",
}

_UNSUPPORTED_BUSINESS_KEYWORDS = ("转换", "分红", "红利", "折算", "拆分", "冻结", "撤销", "撤单")
_CONFIRMED_STATUS_KEYWORDS = ("确认成功", "交易成功", "已确认", "已完成", "成功", "完成")

_ASSET_TYPE_MAP = {
    "基金": "fund",
    "fund": "fund",
    "股票": "stock",
    "stock": "stock",
}


def _normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    return re.sub(r"[\s_\-()/\\\[\]{}:：]+", "", text)


def _decode_csv(data: bytes) -> tuple[str, str]:
    if not data:
        raise ValueError("账单文件为空")
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise ValueError("账单编码无法识别，请导出 UTF-8 或 GBK/GB18030 格式的 CSV")


def _read_csv(data: bytes) -> tuple[list[str], list[dict[str, Any]], str, str]:
    text, encoding = _decode_csv(data)
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [str(header or "").strip() for header in (reader.fieldnames or [])]
    if not headers:
        raise ValueError("未识别到账单表头，请确认文件为带表头的 CSV")
    rows = []
    for row in reader:
        cleaned = {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
        if any(cleaned.values()):
            rows.append(cleaned)
        if len(rows) > _MAX_ROWS:
            raise ValueError(f"单次最多预览 {_MAX_ROWS} 条交易，请拆分账单后导入")
    delimiter_label = {",": "逗号", ";": "分号", "\t": "制表符"}.get(delimiter, delimiter)
    return headers, rows, f"CSV · {delimiter_label}", encoding


def _read_xlsx(data: bytes) -> tuple[list[str], list[dict[str, Any]], str, str]:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Excel 交易账单无法读取: {exc}") from exc

    try:
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        header_values = next(values, None)
        while header_values is not None and not any(str(value or "").strip() for value in header_values):
            header_values = next(values, None)
        headers = [str(value or "").strip() for value in (header_values or [])]
        if not headers or not any(headers):
            raise ValueError("未识别到 Excel 交易账单表头")

        rows = []
        for row_values in values:
            row = {
                headers[index]: row_values[index] if index < len(row_values) else ""
                for index in range(len(headers))
                if headers[index]
            }
            if any(str(value or "").strip() for value in row.values()):
                rows.append(row)
            if len(rows) > _MAX_ROWS:
                raise ValueError(f"单次最多预览 {_MAX_ROWS} 条交易，请拆分账单后导入")
        return headers, rows, "Excel", "xlsx"
    finally:
        workbook.close()


def _read_statement(data: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]], str, str]:
    suffix = str(filename or "").lower().rsplit(".", 1)[-1]
    if suffix == "csv":
        return _read_csv(data)
    if suffix == "xlsx":
        return _read_xlsx(data)
    raise ValueError("请上传 CSV 或 XLSX 格式的交易账单")


def _find_mapping(headers: list[str]) -> dict[str, str | None]:
    normalized_headers = {_normalize(header): header for header in headers if header}
    mapping = {}
    for field, aliases in _HEADER_ALIASES.items():
        mapping[field] = next((normalized_headers.get(_normalize(alias)) for alias in aliases if _normalize(alias) in normalized_headers), None)
    return mapping


def _value(row: dict[str, Any], mapping: dict[str, str | None], field: str) -> Any:
    header = mapping.get(field)
    return row.get(header) if header else ""


def _number(value: Any) -> float | None:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    if not text or text in {"-", "--", "N/A", "n/a"}:
        return None
    text = text.replace(",", "").replace("¥", "").replace("￥", "").replace("元", "")
    text = re.sub(r"\s+", "", text)
    try:
        return float(text)
    except ValueError:
        return None


def _clean_code(value: Any, *, pad_numeric_code: bool = False) -> str:
    code = unicodedata.normalize("NFKC", str(value or "")).strip()
    if re.fullmatch(r"\d+\.0", code):
        code = code[:-2]
    if pad_numeric_code and code.isdigit() and len(code) < 6:
        return code.zfill(6)
    return code


def _parse_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    if not text:
        return None
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace("/", "-").replace(".", "-")
    text = re.sub(r"\s+.*$", "", text)
    for pattern in ("%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return None


def _map_trade_type(value: Any) -> str | None:
    normalized = _normalize(value)
    if not normalized or any(_normalize(keyword) in normalized for keyword in _UNSUPPORTED_BUSINESS_KEYWORDS):
        return None
    for label, result in _TYPE_MAP.items():
        if _normalize(label) == normalized:
            return result
    if "申购" in normalized or "认购" in normalized or "定投" in normalized:
        return "buy"
    if "赎回" in normalized:
        return "sell"
    return None


def _map_asset_type(value: str, default_asset_type: str) -> str:
    return _ASSET_TYPE_MAP.get(_normalize(value), default_asset_type)


def _is_confirmed_status(value: Any) -> bool:
    normalized = _normalize(value)
    if not normalized:
        return True
    return any(_normalize(keyword) in normalized for keyword in _CONFIRMED_STATUS_KEYWORDS)


def _text(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def _detect_template(filename: str, headers: list[str]) -> dict[str, str]:
    normalized_headers = {_normalize(header) for header in headers if header}
    tiantian_signals = {_normalize(value) for value in ("确认日期", "确认份额", "确认净值", "基金代码", "业务名称")}
    filename_text = _normalize(filename)
    if "天天基金" in str(filename or "") or len(normalized_headers & tiantian_signals) >= 3:
        return {
            "id": "tiantian_fund_transaction",
            "label": "天天基金确认交易流水",
            "source": "tiantian_fund_transaction_export",
            "detection": "确认字段与基金业务列",
        }
    return {
        "id": "transaction_statement",
        "label": "交易流水账单",
        "source": "csv_import",
        "detection": "通用字段映射",
    }


def parse_transaction_file(
    data: bytes,
    *,
    filename: str = "",
    default_asset_type: str = "fund",
    default_market: str = "基金",
) -> dict:
    """Build an editable transaction preview without writing the uploaded file."""
    default_asset_type = str(default_asset_type or "").strip()
    if default_asset_type not in {"fund", "stock"}:
        raise ValueError("默认资产类型必须是基金或股票")
    if not data:
        raise ValueError("账单文件为空")

    headers, rows, file_format, encoding = _read_statement(data, filename)
    mapping = _find_mapping(headers)
    missing_columns = [field for field in ("trade_date", "code", "trade_type", "shares") if not mapping.get(field)]
    if missing_columns:
        labels = {"trade_date": "确认/交易日期", "code": "代码", "trade_type": "交易方向", "shares": "确认份额/数量"}
        raise ValueError(f"账单缺少必要列: {'、'.join(labels[field] for field in missing_columns)}")
    if not mapping.get("unit_price") and not mapping.get("gross_amount"):
        raise ValueError("账单至少需要确认/成交单价或确认/成交金额其中一列")

    template = _detect_template(filename, headers)
    candidates = []
    errors = []
    warnings = []
    for index, row in enumerate(rows, start=2):
        raw_date = _value(row, mapping, "trade_date")
        asset_type = _map_asset_type(_text(_value(row, mapping, "asset_type")), default_asset_type)
        raw_code = _text(_value(row, mapping, "code"))
        code = _clean_code(raw_code, pad_numeric_code=asset_type == "fund")
        raw_type = _text(_value(row, mapping, "trade_type"))
        status = _text(_value(row, mapping, "status"))
        shares = _number(_value(row, mapping, "shares"))
        unit_price = _number(_value(row, mapping, "unit_price"))
        gross_amount = _number(_value(row, mapping, "gross_amount"))
        fee = _number(_value(row, mapping, "fee")) or 0
        trade_date = _parse_date(raw_date)
        trade_type = _map_trade_type(raw_type)

        row_errors = []
        if not trade_date:
            row_errors.append("交易日期无法识别")
        if not code:
            row_errors.append("代码为空")
        if not trade_type:
            row_errors.append(f"交易方向无法识别或现金流含义不完整: {raw_type or '-'}")
        if mapping.get("status") and not _is_confirmed_status(status):
            row_errors.append(f"交易状态不是已确认成功: {status or '-'}")
        if shares is None or shares <= 0:
            row_errors.append("成交份额必须大于 0")
        if unit_price is None and gross_amount is not None and shares and shares > 0:
            unit_price = gross_amount / shares
            warnings.append(f"第 {index} 行未提供确认/成交单价，已由金额除以份额计算；请确认。")
        if unit_price is None or unit_price <= 0:
            row_errors.append("成交单价必须大于 0")
        if fee < 0:
            row_errors.append("费用不能小于 0")
        if row_errors:
            errors.append({"row": index, "message": "；".join(row_errors)})
            continue

        if raw_code != code and code:
            warnings.append(f"第 {index} 行基金代码已补齐前导零为 {code}；请确认。")
        market = _text(_value(row, mapping, "market")) or str(default_market or "").strip()
        note = _text(_value(row, mapping, "note"))
        if not note and raw_type:
            note = f"导入业务: {raw_type}"
        candidates.append({
            "asset_type": asset_type,
            "market": market,
            "code": code,
            "name": _value(row, mapping, "name"),
            "trade_type": trade_type,
            "trade_date": trade_date,
            "shares": round(shares, 8),
            "unit_price": round(unit_price, 8),
            "fee": round(fee, 8),
            "note": note,
            "source": template["source"],
            "csv_row": index,
        })

    if not candidates:
        raise ValueError("账单中没有可导入的有效交易，请处理预览错误后重试")
    if errors:
        warnings.append(f"有 {len(errors)} 行无法导入，已从待确认列表中排除。")
    return {
        "source": "用户上传交易账单（仅预览，未写入账本）",
        "filename": str(filename or "")[:255],
        "file_sha256": hashlib.sha256(data).hexdigest(),
        "encoding": encoding,
        "delimiter": file_format,
        "format": file_format,
        "template": template,
        "headers": headers,
        "mapping": mapping,
        "candidates": candidates,
        "errors": errors,
        "warnings": warnings,
        "privacy": "原始 CSV 不会保存；Excel 交易账单同样不会保存。确认导入后仅保存交易字段、来源和文件哈希，用于防止重复导入。",
    }


def parse_transaction_csv(
    data: bytes,
    *,
    filename: str = "",
    default_asset_type: str = "fund",
    default_market: str = "基金",
) -> dict:
    """Backward-compatible name for callers that still use the CSV endpoint."""
    return parse_transaction_file(
        data,
        filename=filename,
        default_asset_type=default_asset_type,
        default_market=default_market,
    )
